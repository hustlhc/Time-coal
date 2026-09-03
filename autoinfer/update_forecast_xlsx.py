import json
import openpyxl
from datetime import datetime
import os
import csv

# 配置文件路径
JSON_DIR = 'json'  # JSON文件存放目录
CSV_FILE = '../dataset/pre_coal/coal_new.csv'  # 真实值CSV文件路径
EXCEL_FILE = '预测对比.xlsx'  # Excel文件路径

# 查找最新的JSON文件
def find_latest_json(json_dir):
    """从指定目录查找最新的JSON文件"""
    json_files = []
    
    # 确保目录存在
    if not os.path.exists(json_dir):
        raise FileNotFoundError(f"JSON文件目录 {json_dir} 不存在")
    
    for file in os.listdir(json_dir):
        if file.endswith('_data.json'):
            try:
                # 提取日期部分
                date_str = file.split('_')[0]
                date = datetime.strptime(date_str, '%Y-%m-%d')
                json_files.append((date, os.path.join(json_dir, file)))
            except ValueError:
                # 忽略不符合日期格式的文件
                pass
    
    if not json_files:
        raise FileNotFoundError(f"在 {json_dir} 目录中未找到符合格式的JSON文件")
    
    # 按日期排序，返回最新的文件
    json_files.sort(key=lambda x: x[0], reverse=True)
    return json_files[0][1]

# 从CSV文件读取真实值
def load_real_values(csv_file):
    """从CSV文件读取真实值，真实值在倒数第三列"""
    real_values = {}
    
    if not os.path.exists(csv_file):
        raise FileNotFoundError(f"CSV文件 {csv_file} 不存在")
    
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        header = next(reader)  # 读取表头
        
        # 计算倒数第三列的索引
        real_value_col_index = len(header) - 3
        
        for row in reader:
            if row:
                date_str = row[0]  # 第一列是日期
                try:
                    # 解析日期格式
                    datetime.strptime(date_str, '%Y-%m-%d')
                    # 读取倒数第三列的真实值
                    real_value = float(row[real_value_col_index])
                    real_values[date_str] = real_value
                except (ValueError, IndexError):
                    # 忽略格式错误的行
                    pass
    
    return real_values

# 读取最新的JSON文件
data_file = find_latest_json(JSON_DIR)
print(f"正在读取最新的JSON文件: {data_file}")
with open(data_file, 'r', encoding='utf-8') as f:
    data = json.load(f)

# 提取CCI3800的预测数据
cci3800_data = data['data']['CCI3800outinfer']

# 获取最新一天的日期
latest_date = data['inferDate']
print(f"最新预测日期: {latest_date}")

# 将预测数据转换为字典，键为日期，值为预测价格
forecast_dict = {item['date']: item['predict'] for item in cci3800_data}

# 读取真实值
print(f"正在读取CSV文件: {CSV_FILE}")
real_values = load_real_values(CSV_FILE)
print(f"成功读取 {len(real_values)} 条真实值数据")

# 使用openpyxl打开Excel文件，保留原有格式
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active

# 查找日期列、每周修正预测价格列和真实值列的位置
date_col = 1  # 第一列始终是日期列
forecast_col = None
real_col = None

# 遍历第一行查找相关列
for col in range(1, ws.max_column + 1):
    cell_value = ws.cell(row=1, column=col).value
    if cell_value == '每周修正预测价格':
        forecast_col = col
    elif cell_value == 'CCI价格指数':
        real_col = col

if forecast_col is None:
    print("错误：未找到每周修正预测价格列")
    exit(1)

# 如果真实值列不存在，在第二列创建新列
if real_col is None:
    real_col = 2  # 真实值列固定在第二列
    ws.cell(row=1, column=real_col).value = 'CCI价格指数'
    print(f"未找到真实值列，在第{real_col}列创建了新的CCI价格指数列")
else:
    print(f"找到日期列在第{date_col}列，每周修正预测价格列在第{forecast_col}列，真实值列在第{real_col}列")

# 只更新最新一天及其之后的预测数据和真实值
updated_count = 0
for row in range(2, ws.max_row + 1):
    date_cell = ws.cell(row=row, column=date_col)
    date_value = date_cell.value
    
    # 处理日期格式
    if isinstance(date_value, datetime):
        date_str = date_value.strftime('%Y-%m-%d')
    else:
        date_str = str(date_value).strip()
    
    # 只更新最新一天及其之后的日期
    if date_str >= latest_date:
        # 更新预测值
        if date_str in forecast_dict:
            # 获取表头的格式作为标准格式
            header_format = ws.cell(row=1, column=forecast_col).number_format
            # 写入新值
            ws.cell(row=row, column=forecast_col).value = forecast_dict[date_str]
            # 设置为标准格式以显示为整数
            ws.cell(row=row, column=forecast_col).number_format = header_format
        
        # 更新真实值
        if date_str in real_values:
            # 获取真实值列的格式
            real_header_format = ws.cell(row=1, column=real_col).number_format
            # 写入真实值
            ws.cell(row=row, column=real_col).value = real_values[date_str]
            # 设置格式
            ws.cell(row=row, column=real_col).number_format = real_header_format
        
        updated_count += 1

# 保存更新后的Excel文件
wb.save(EXCEL_FILE)

print(f"已成功更新 {EXCEL_FILE} 文件中的每周修正预测价格列和真实值列")
print(f"只更新了 {latest_date} 及其之后的 {updated_count} 条记录")
print(f"成功读取 {len(real_values)} 条真实值数据")
print("保留了原Excel文件的显示格式")

